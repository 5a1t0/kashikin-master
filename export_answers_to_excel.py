import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# データベースに接続
conn = sqlite3.connect('kashikin.db')
cursor = conn.cursor()

# Excelワークブックを作成
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "解答一覧"

# ヘッダー行のスタイル
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

# ヘッダーを作成
ws['A1'] = '問題番号'
ws['B1'] = '令和5年度'
ws['C1'] = '令和4年度'
ws['D1'] = '令和3年度'
ws['E1'] = '令和2年度'
ws['F1'] = '令和元年度'

for col in ['A', 'B', 'C', 'D', 'E', 'F']:
    ws[f'{col}1'].fill = header_fill
    ws[f'{col}1'].font = header_font
    ws[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center')

# 各年度の解答を取得
years = ['令和5年度', '令和4年度', '令和3年度', '令和2年度', '令和元年度']
year_answers = {}

for year in years:
    cursor.execute("""
        SELECT answer 
        FROM questions 
        WHERE year = ? 
        ORDER BY id
    """, (year,))
    
    answers = cursor.fetchall()
    year_answers[year] = [row[0] for row in answers]

# データを行ごとに書き込み
max_questions = max(len(answers) for answers in year_answers.values())

for i in range(max_questions):
    row = i + 2
    ws[f'A{row}'] = f'第{i+1}問'
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    for col_idx, year in enumerate(years, start=2):
        col_letter = chr(65 + col_idx)  # B, C, D, E, F
        if i < len(year_answers[year]):
            ws[f'{col_letter}{row}'] = year_answers[year][i]
            ws[f'{col_letter}{row}'].alignment = Alignment(horizontal='center')

# 列幅を調整
ws.column_dimensions['A'].width = 12
for col in ['B', 'C', 'D', 'E', 'F']:
    ws.column_dimensions[col].width = 15

# ファイルを保存
filename = '解答一覧_令和5年度～令和元年度.xlsx'
wb.save(filename)
conn.close()

print(f"✅ Excelファイルを作成しました: {filename}")
print(f"📊 合計: {max_questions}問 × 5年度")
